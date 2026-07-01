"""
Comparator dùng để so sánh 2 file Excel (.xlsx).

Cách hoạt động: tương tự CsvComparator, nhưng đọc bằng openpyxl để giữ
đúng định dạng số/text của Excel, và hỗ trợ so sánh theo từng sheet
(vì 1 file Excel có thể có nhiều sheet, CSV thì không).
"""

import os

import openpyxl

from app.core.comparators.base_comparator import BaseComparator
from app.core.models.diff_result import ChangeType, DiffItem, DiffResult


class ExcelComparator(BaseComparator):
    """So sánh 2 file Excel theo từng sheet, từng dòng, từng cột."""

    def validate_files(self, file_a: str, file_b: str) -> tuple[bool, str]:
        for path in (file_a, file_b):
            if not os.path.exists(path):
                return False, f"Không tìm thấy file: {path}"
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return False, f"File không đúng định dạng Excel (.xlsx): {path}"
        return True, ""

    def compare(self, file_a: str, file_b: str) -> DiffResult:
        result = DiffResult(file_a=file_a, file_b=file_b)

        # data_only=True: lấy giá trị đã tính toán của công thức, không lấy công thức thô
        wb_a = openpyxl.load_workbook(file_a, data_only=True)
        wb_b = openpyxl.load_workbook(file_b, data_only=True)

        # So sánh danh sách sheet trước
        sheets_a = set(wb_a.sheetnames)
        sheets_b = set(wb_b.sheetnames)

        for sheet_name in sheets_a - sheets_b:
            result.items.append(
                DiffItem(
                    location=f"Sheet '{sheet_name}'",
                    change_type=ChangeType.REMOVED,
                    old_value="(toàn bộ sheet)",
                )
            )

        for sheet_name in sheets_b - sheets_a:
            result.items.append(
                DiffItem(
                    location=f"Sheet '{sheet_name}'",
                    change_type=ChangeType.ADDED,
                    new_value="(toàn bộ sheet)",
                )
            )

        # So sánh nội dung các sheet có mặt ở cả 2 file
        for sheet_name in sheets_a & sheets_b:
            self._compare_sheet(wb_a[sheet_name], wb_b[sheet_name], sheet_name, result)

        return result

    def _compare_sheet(self, ws_a, ws_b, sheet_name: str, result: DiffResult) -> None:
        """So sánh nội dung 2 sheet cùng tên, từng ô một theo tọa độ (dòng, cột)."""
        max_row = max(ws_a.max_row, ws_b.max_row)
        max_col = max(ws_a.max_column, ws_b.max_column)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val_a = ws_a.cell(row=row, column=col).value
                val_b = ws_b.cell(row=row, column=col).value

                if val_a == val_b:
                    continue

                cell_ref = ws_a.cell(row=row, column=col).coordinate
                location = f"Sheet '{sheet_name}', ô {cell_ref}"

                if val_a is None and val_b is not None:
                    change_type = ChangeType.ADDED
                elif val_a is not None and val_b is None:
                    change_type = ChangeType.REMOVED
                else:
                    change_type = ChangeType.MODIFIED

                result.items.append(
                    DiffItem(
                        location=location,
                        change_type=change_type,
                        old_value=None if val_a is None else str(val_a),
                        new_value=None if val_b is None else str(val_b),
                    )
                )
