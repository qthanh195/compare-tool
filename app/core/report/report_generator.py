"""
Module xuất báo cáo kết quả so sánh ra file Excel, có tô màu theo loại thay đổi.

Dùng chung cho mọi comparator (Excel/CSV/DXF...) vì đầu vào luôn là
DiffResult đã được chuẩn hóa — đây chính là lợi ích của việc chuẩn hóa
cấu trúc dữ liệu kết quả ngay từ đầu.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

from app.core.models.diff_result import ChangeType, DiffResult

# Màu nền cho từng loại thay đổi khi xuất báo cáo
_COLOR_BY_CHANGE_TYPE = {
    ChangeType.MODIFIED: "FFF2CC",  # vàng nhạt
    ChangeType.ADDED: "D9EAD3",  # xanh lá nhạt
    ChangeType.REMOVED: "F4CCCC",  # đỏ nhạt
}

_LABEL_BY_CHANGE_TYPE = {
    ChangeType.MODIFIED: "Thay đổi",
    ChangeType.ADDED: "Thêm mới",
    ChangeType.REMOVED: "Đã xóa",
}


def export_diff_report(result: DiffResult, output_path: str) -> str:
    """
    Xuất DiffResult ra file Excel báo cáo, có tô màu theo loại thay đổi.

    Trả về đường dẫn file báo cáo đã tạo (output_path), để UI hiển thị
    thông báo "Đã lưu báo cáo tại: ..." hoặc mở file luôn cho người dùng.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo so sánh"

    # Dòng tóm tắt trên cùng
    summary = result.summary
    ws.append([
        f"So sánh: {result.file_a}  <>  {result.file_b}"
    ])
    ws.append([
        f"Tổng số thay đổi: {len(result.items)}  "
        f"(Thay đổi: {summary['modified']}, "
        f"Thêm mới: {summary['added']}, "
        f"Đã xóa: {summary['removed']})"
    ])
    ws.append([])  # dòng trống

    # Header bảng chi tiết
    header = ["Vị trí", "Loại thay đổi", "Giá trị cũ", "Giá trị mới"]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Chi tiết từng thay đổi, tô màu theo loại
    for item in result.items:
        ws.append([
            item.location,
            _LABEL_BY_CHANGE_TYPE[item.change_type],
            item.old_value or "",
            item.new_value or "",
        ])
        fill = PatternFill(
            start_color=_COLOR_BY_CHANGE_TYPE[item.change_type],
            end_color=_COLOR_BY_CHANGE_TYPE[item.change_type],
            fill_type="solid",
        )
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # Tự động giãn độ rộng cột cho dễ đọc
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

    wb.save(output_path)
    return output_path
